# -*- coding: utf-8 -*-
"""
Created on Mon Jan 14 14:27:07 2019

@author: Kraan
"""
from openpyxl import load_workbook
from collections import Counter
import json
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
stopwords = stopwords.words('Dutch')

sources = [['teksten op de fotos bronckhorst energierallys.xlsx', 'tekst'],
           ["Bezetting + aanmelding Energierally's.xlsx", 'open vragen'],
           ['13011_Dataset onderzoek Duurzaamheid.xlsx', 'enquette']]


def countrows(ws, colnum=1, rownum=1):
    column = colnum
    row = rownum
    value = ws.cell(row=row, column=column).value
    while value != '' and value is not None:
        row += 1
        value = ws.cell(row=row, column=column).value
    return(row-1)


def countcols(ws, colnum=1, rownum=1):
    column = colnum
    row = rownum
    value = ws.cell(row=row, column=column).value
    while value != '' and value is not None:
        column += 1
        value = ws.cell(row=row, column=column).value
    return(column-1)


def cleantext(inputtext):
    wordslist = word_tokenize(inputtext.lower())
    for word in wordslist:
        if word not in stopwords and word not in ['.', ',', ':', ';']:
            yield word


def getwordcount(inputlist):
    text = ''
    for textpart in inputlist:
        text += textpart + ' '
    words = cleantext(text)
    return(Counter(words))


totalresult = []
for source in sources:
    wb = load_workbook(filename='data/' + source[0])
    if 'dataprep' in wb.sheetnames:
        ws = wb['dataprep']
    else:
        ws = wb.active

    rows = countrows(ws=ws, colnum=1, rownum=4)
    columns = countcols(ws=ws, colnum=1, rownum=1)

    if source[1] == 'enquette':
        result = []
        for col in range(2, columns):
            question = ws.cell(row=2, column=col).value.lower()
            questype = ws.cell(row=3, column=col).value.lower()
            subresult = []
            for row in range(4, rows):
                cellvalue = ws.cell(row=row, column=col).value
                if cellvalue == 0 or cellvalue is None or cellvalue == '':
                    if questype == 'meerkeuze':
                        cellvalue = 'niet aangevinkt'
                        subresult.append(cellvalue.lower().strip())
                elif type(cellvalue) == str:
                    subresult.append(cellvalue.lower().strip())
                else:
                    subresult.append(cellvalue)
            if questype == 'meerkeuze':
                subresult = [Counter(subresult)]
                wordcount = ''
            elif questype == 'open':
                wordcount = getwordcount(subresult)
            subresult = [question, questype, subresult, col-1, wordcount]
            result.append(subresult)

    elif source[1] == 'open vragen':
        result = []
        for row in range(2, rows):
            question = ws.cell(row=row, column=1).value.lower()
            questype = 'open'
            subresult = []
            for col in range(2, columns):
                cellvalue = ws.cell(row=row, column=col).value
                if cellvalue is not None and cellvalue != 0 and cellvalue != '':
                    subresult.append(cellvalue.lower())
            wordcount = getwordcount(subresult)
            subresult = [question, questype, subresult, row-1, wordcount]
            result.append(subresult)

    elif source[1] == 'tekst':
        subresult = []
        question = 'Verzamelde teksten'
        questype = 'open'
        for row in range(2, rows):
            subresult.append(ws.cell(row=row, column=1).value.lower())
        wordcount = getwordcount(subresult)
        result = [[question, questype, subresult, 1, wordcount]]

    for row in result:
        totalresult.append({'author': 'inwoner',
                            'date': '1-1-2019',
                            'document': row[2],
                            'id': source[0]+'_'+str(row[3]),
                            'masterID': source[0],
                            'place': 'Brockhorst',
                            'summary': row[0],
                            'sourcetype': source[1],
                            'questype': row[1],
                            'wordcounter': row[4]})
# save data
with open("xtra_data.json", "w") as f:
    json.dump(totalresult, f)
